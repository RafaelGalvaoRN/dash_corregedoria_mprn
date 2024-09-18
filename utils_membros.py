import streamlit as st
from reportlab.pdfgen import canvas
import zipfile

from PyPDF2 import PdfReader

from relacao_promotorias import promotoria, promotores
from datetime import datetime
import io

from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

import pandas as pd
from datetime import date
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle



def append_to_excel_manually(filename, df_membros, df_substituido, sheet_name='Sheet1'):
    # Tentar abrir o arquivo existente, ou criar um novo se não existir
    try:
        workbook = load_workbook(filename)
        sheet = workbook[sheet_name]
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

    # Encontrar a última linha preenchida na planilha para continuar adicionando dados
    startrow = sheet.max_row + 1

    # Adicionar os dados de df_membros
    sheet.append(["Promotoria Selecionada", "Promotor", "Antiguidade", "Data Selecionada",
                  "Órgão Ministerial da Última Correicão", "Registro de Pena"])
    for row in dataframe_to_rows(df_membros, index=False, header=False):
        sheet.append(row)

    # Adicionar um espaçamento entre as tabelas
    startrow = sheet.max_row + 2
    sheet.append([])  # Linha em branco para separar

    # Adicionar os dados de df_substituido
    sheet.append(
        ["Informações", "Localização das Informações", "Informação Conceito ou Registro Disciplinar", "Observações"])
    for row in dataframe_to_rows(df_substituido, index=False, header=False):
        sheet.append(row)

    # Salvar o arquivo atualizado
    workbook.save(filename)


# Função para salvar e baixar o arquivo Excel gerado manualmente
# def append_to_excel_manually_and_download(df_membros, df_substituido, nome_arquivo, sheet_name='Sheet1'):
#     # Criar um buffer de bytes para segurar o arquivo Excel
#     output = io.BytesIO()
#
#     # Tentar abrir o arquivo existente, ou criar um novo se não existir
#     workbook = Workbook()
#     sheet = workbook.active
#     sheet.title = sheet_name
#
#     # Adicionar os dados de df_membros
#     sheet.append(df_membros.columns.tolist())
#     for row in dataframe_to_rows(df_membros, index=False, header=False):
#         sheet.append(row)
#
#     # Adicionar um espaçamento entre as tabelas
#     sheet.append([])  # Linha em branco para separar
#
#     # Adicionar os dados de df_substituido
#     sheet.append(df_substituido.columns.tolist())
#     for row in dataframe_to_rows(df_substituido, index=False, header=False):
#         sheet.append(row)
#
#     # Salvar o arquivo no buffer de bytes
#     workbook.save(output)
#
#     # Voltar para o início do buffer
#     output.seek(0)
#
#
#     file_name_saida = f"{nome_arquivo}_dados_tratados.xlsx"
#     workbook.save(file_name_saida)
#
#
#     # Criar o botão de download para o arquivo Excel
#     st.download_button(label="Download dados tratados como Excel",
#                        data=output,
#                        file_name=file_name_saida,
#                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#
#     return file_name_saida

def append_to_excel_manually(df_membros, df_substituido, nome_arquivo, sheet_name='Sheet1'):
    # Criar um novo workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    # Adicionar os dados de df_membros
    sheet.append(df_membros.columns.tolist())
    for row in dataframe_to_rows(df_membros, index=False, header=False):
        sheet.append(row)

    # Adicionar um espaçamento entre as tabelas
    sheet.append([])  # Linha em branco para separar

    # Adicionar os dados de df_substituido
    sheet.append(df_substituido.columns.tolist())
    for row in dataframe_to_rows(df_substituido, index=False, header=False):
        sheet.append(row)

    # Salvar o arquivo no disco
    file_name_saida = f"{nome_arquivo}_dados_tratados.xlsx"
    workbook.save(file_name_saida)

    # Retornar o nome do arquivo salvo
    return file_name_saida


def membros_menu():

    promotoria_ordenada = sorted(promotoria)

    promotoria_selecionada = st.selectbox("Escolha o cargo do Candidato", promotoria_ordenada, key="promotoria_tab5")

    promotor = st.selectbox('Escolha o nome do Membro:', promotores)

    antiguidade = st.number_input("Digite a posição na lista de antiguidade", min_value=1)

    data_selecionada = st.date_input('Data da Última Correição Ordinária', datetime.now().date())

    data_selecionada = data_selecionada.strftime("%d-%m-%Y")

    orgao_ministerial_ultima_correicao = st.selectbox("Escolha o órgão Mnisterial da Última Correição do  Candidato", promotoria_ordenada)


    return (
        promotoria_selecionada, promotor, antiguidade, data_selecionada, orgao_ministerial_ultima_correicao,
        )


def membros_pdf_extract():
    arquivos = st.file_uploader("Junte os PDF's aqui", type="PDF", accept_multiple_files=True)
    return arquivos


def get_page(texto_page: str, string: str) -> bool:
    if string.lower() in texto_page:
        return True
    return False





def get_value_by_partial_key(dictionary, partial_key):
    for key in dictionary:
        if partial_key in key:
            return dictionary[key]
    return None  # Retorna None se a chave parcial não for encontrada


def get_pages(pdf) -> int:
    reader = PdfReader(pdf)

    return len(reader.pages)


# def extract_pdf(pdf, lista_verificacao: list) -> dict:
#     """
#     extrai texto de pdf e verifica pagina inicial e pagina final do relatório procurado
#     :param pdf: texto a ser extraído
#     :param lista_verificacao: lsita com tupla, contendo o tipo da página (inicial ou final) e o texto contido na
#     página
#     :return: dict contendo os dados encontrados
#     """
#
#     # Criando um objeto PdfReader
#     reader = PdfReader(pdf)
#
#     # Inicializando uma variável para armazenar o texto extraído
#     full_text = ""
#
#     result = {}
#
#     # Iterando sobre todas as páginas do PDF
#     for i in range(len(reader.pages)):
#         # Obtendo uma página específica do arquivo PDF
#         page = reader.pages[i]
#
#         # Extraindo o texto da página
#         text = page.extract_text().lower()
#
#         for item in lista_verificacao:
#             if get_page(text, item[1]):
#                 result[item[0]] = i + 1
#
#     return result

import re

# def extract_pdf(pdf, lista_verificacao: list, first_search: bool = True, opcional_string: str = None) -> dict:
#     """
#     Extracts text from a PDF and identifies the initial and final pages of the desired sections.
#
#     :param pdf: Path to the PDF file.
#     :param lista_verificacao: List of tuples, each containing the type of page ('pag_inicial' or 'pag_final') and the text to search for.
#     :param first_search: If False, skips the first occurrence of the initial string.
#     :param opcional_string: Optional string to search for after the final page string.
#     :return: Dictionary containing the found page numbers.
#     """
#
#     print("Inicializando a funcao interna")
#     print("argumento: ", lista_verificacao)
#
#
#     reader = PdfReader(pdf)
#     result = {'pag_inicial': None, 'pag_final': None}
#
#     total_pages = len(reader.pages)
#     found_pag_inicial = False
#     found_pag_final = False
#     initial_string_occurrences = 0  # Counter for occurrences of the initial string
#
#     for i in range(total_pages):
#         page = reader.pages[i]
#         text = page.extract_text()
#         if text:
#             text_lower = text.lower()
#
#             # Search for the initial page string
#             if not found_pag_inicial:
#                 initial_string = lista_verificacao[0][1].lower()
#                 if initial_string in text_lower:
#                     initial_string_occurrences += 1
#                     if not first_search and initial_string_occurrences == 1:
#                         # Skip the first occurrence
#                         continue
#                     result['pag_inicial'] = i + 1  # Pages are 1-indexed
#                     found_pag_inicial = True
#                     continue  # Proceed to next page
#
#             # Search for the final page string
#             if found_pag_inicial and not found_pag_final:
#                 final_string = lista_verificacao[1][1].lower()
#
#                 print("imprimindo argumento lista verificacao", lista_verificacao)
#                 print("verificando pagina final")
#                 print("procurado final string:", final_string)
#                 print("procurando no texto: ",text_lower )
#
#                 if final_string in text_lower:
#                     result['pag_final'] = i + 1
#                     found_pag_final = True
#
#                     # If an optional string is provided, search for it after finding the final page string
#                     if opcional_string:
#                         opcional_string_lower = opcional_string.lower()
#                         # Continue searching from the next page
#                         for j in range(i + 1, total_pages):
#                             next_page = reader.pages[j]
#                             next_text = next_page.extract_text()
#                             if next_text and opcional_string_lower in next_text.lower():
#                                 result['pag_final'] = j + 1
#                                 break
#                     break  # Stop processing after finding the final page (and optional string if provided)
#
#     print("imprimindo retorno da funcao", result)
#     return result



def extract_pdf(pdf, lista_verificacao: list, first_search: bool = True, opcional_string: str = None) -> dict:
    """
    Extracts text from a PDF and identifies the initial and final pages of the desired sections.
    Optionally, it searches for an additional string after finding the initial final page.

    :param pdf: Path to the PDF file.
    :param lista_verificacao: List of tuples, each containing the type of page ('pag_inicial' or 'pag_final') and the text to search for.
    :param first_search: If False, skips the first occurrence of the initial string.
    :param opcional_string: Optional string to search for after the final page string is found.
    :return: Dictionary containing the found page numbers.
    """
    # print("Initializing function")
    # print("Arguments:", lista_verificacao)

    reader = PdfReader(pdf)
    result = {'pag_inicial': None, 'pag_final': None, 'opcional_string': None}

    total_pages = len(reader.pages)
    initial_string_occurrences = 0  # Counter for occurrences of the initial string

    # Step 1: Search for 'pag_inicial'
    for i in range(total_pages):
        page = reader.pages[i]
        text = page.extract_text()
        if text:
            text_lower = text.lower()
            initial_string = lista_verificacao[0][1].lower()
            if initial_string in text_lower:
                initial_string_occurrences += 1
                if not first_search and initial_string_occurrences == 1:
                    # print(f"Skipping first occurrence of initial string on page {i + 1}")
                    pass  # Skips the first occurrence if first_search is False
                else:
                    result['pag_inicial'] = i + 1  # Pages are 1-indexed
                    # print(f"Found 'pag_inicial' on page {i + 1}")
                    break  # Stop searching for 'pag_inicial'

    # Check if 'pag_inicial' is None
    if result['pag_inicial'] is None:
        # print("Error: 'pag_inicial' not found.")
        return result

    # Step 2: Search for 'pag_final' based on the second string in lista_verificacao
    final_string = lista_verificacao[1][1].lower()
    if result['pag_inicial'] is not None:
        # print("Searching for 'pag_final'")
        start_index = result['pag_inicial'] - 1  # Convert to zero-based index
        for i in range(start_index, total_pages):
            page = reader.pages[i]
            text = page.extract_text()
            if text:
                text_lower = text.lower()
                if final_string in text_lower:
                    result['pag_final'] = i + 1  # Pages are 1-indexed
                    # print(f"Found 'pag_final' on page {i + 1}")
                    break  # Stop searching after finding 'pag_final'

    # Step 3: If opcional_string is provided, search for it after 'pag_final'
    if opcional_string and result['pag_final'] is not None:
        # print("Searching for optional string:", opcional_string)
        opcional_string_lower = opcional_string.lower()
        start_index = result['pag_final'] - 1  # Convert to zero-based index (starting from pag_final)
        for i in range(start_index, total_pages):
            page = reader.pages[i]
            text = page.extract_text()
            if text:
                text_lower = text.lower()
                if opcional_string_lower in text_lower:
                    result['opcional_string'] = i + 1  # Pages are 1-indexed
                    # print(f"Found 'opcional_string' on page {i + 1}")
                    break  # Stop searching once optional string is found

    # print("Result:", result)
    return result

import unicodedata
import string




# Função para substituir placeholders pelos valores correspondentes da lista 'campos_preencher'
def substituir_placeholders(dados, campos_preencher):
    for chave, valores in dados.items():
        for i, valor in enumerate(valores):
            for arquivo, descricao, folhas in campos_preencher:
                if '{{' in valor:  # Verifica se há um placeholder no valor
                    # Substitui o placeholder pelo valor correspondente da lista campos_preencher
                    valor = valor.replace(f'{{{{{descricao}}}}}', folhas)
            dados[chave][i] = valor  # Atualiza o valor substituído no dicionário original
    return dados


def preencher_pdf(output_pdf_path, lista_campos):
    """
    Preenche um PDF com campos específicos (números de páginas, etc.)
    :param output_pdf_path: Caminho para o arquivo de saída (PDF preenchido)
    :param lista_campos: Lista com os textos e coordenadas a serem inseridos no PDF
    """
    # Crie um novo PDF com o texto preenchido nos locais desejados
    c = canvas.Canvas(output_pdf_path, pagesize=letter)

    # Iterando sobre a lista de campos e preenchendo o PDF
    for campo in lista_campos:
        texto, x, y = campo  # Texto a ser preenchido e coordenadas
        c.drawString(x, y, texto)  # Posição X, Y no PDF para colocar o texto

    # Finaliza e salva o PDF
    c.save()


def gerador_indice_pdf(arquivo_excel: str, output_pdf: str):
    # Ler o arquivo Excel em um DataFrame do pandas, preservando as linhas em branco
    df = pd.read_excel(arquivo_excel, header=None, keep_default_na=False, dtype=str)

    # Identificar as linhas em branco (que separarão as duas tabelas)
    empty_row_indices = df.index[df.isnull().all(axis=1) | (df == '').all(axis=1)].tolist()

    if not empty_row_indices:
        raise ValueError("Não foi encontrada uma linha em branco para separar as duas tabelas.")

    # Dividir o DataFrame em duas tabelas
    first_table_df = df.iloc[:empty_row_indices[0]].reset_index(drop=True)
    second_table_df = df.iloc[empty_row_indices[0] + 1:].reset_index(drop=True)

    # Remover possíveis linhas em branco adicionais
    first_table_df = first_table_df.dropna(how='all').replace('', pd.NA).dropna(how='all')
    second_table_df = second_table_df.dropna(how='all').replace('', pd.NA).dropna(how='all')

    # Processar a primeira tabela
    # Definir a primeira linha como cabeçalho
    header_row_first = first_table_df.iloc[0]
    non_null_cols_first = header_row_first.notna() & (header_row_first != '')
    # Selecionar apenas as colunas com nomes válidos
    first_table_df = first_table_df.loc[:, non_null_cols_first]
    header_row_first = header_row_first[non_null_cols_first]
    # Definir os nomes das colunas
    first_table_df.columns = header_row_first
    # Remover a linha de cabeçalho dos dados
    first_table_df = first_table_df[1:].reset_index(drop=True)

    # Processar a segunda tabela
    header_row_second = second_table_df.iloc[0]
    non_null_cols_second = header_row_second.notna() & (header_row_second != '')
    # Selecionar apenas as colunas com nomes válidos
    second_table_df = second_table_df.loc[:, non_null_cols_second]
    header_row_second = header_row_second[non_null_cols_second]
    # Definir os nomes das colunas
    second_table_df.columns = header_row_second
    # Remover a linha de cabeçalho dos dados
    second_table_df = second_table_df[1:].reset_index(drop=True)

    # Criar um SimpleDocTemplate
    doc = SimpleDocTemplate(output_pdf, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    today_date = date.today().strftime("%d/%m/%Y")

    header_texts = ["CORREGEDORIA GERAL DO MPRN", "RELATÓRIO AUTOMATIZADO DE INFORMAÇÕES DO MEMBRO",
                    f"Relatório emitido em {today_date}"]
    header_image = r"img/logo-mprn.png"

    # Cabeçalho
    # Adicionar a imagem do cabeçalho
    img = Image(header_image)
    img.drawHeight = 1.0 * inch
    img.drawWidth = 3.0 * inch
    img.hAlign = 'CENTER'
    elements.append(img)
    elements.append(Spacer(1, 0.2 * inch))

    font_size = 20  # Começa com 14, por exemplo

    # Adicionar as linhas de texto do cabeçalho
    for text in header_texts:
        style_with_dynamic_font = ParagraphStyle(name='DynamicFont', fontSize=font_size,
                                                 alignment=1, leading=font_size + 6)

        paragraph = Paragraph(text, style_with_dynamic_font)
        paragraph.hAlign = 'CENTER'
        elements.append(paragraph)
        elements.append(Spacer(1, 0.3 * inch))

        font_size -= 4  # Diminui o tamanho da fonte em 2 a cada iteração

    # Adicionar espaço após o cabeçalho
    elements.append(Spacer(0.1, 0.01 * inch))

    # Função auxiliar para criar uma tabela a partir de um DataFrame
    def create_table_from_df(df_table, col_widths):
        data = [df_table.columns.tolist()] + df_table.values.tolist()

        # Converter células em Paragraph para suportar quebra de linha
        styleN = styles['Normal']
        for i in range(len(data)):
            for j in range(len(data[i])):
                data[i][j] = Paragraph(str(data[i][j]), styleN)

        # Criar a tabela com os dados e colunas ajustadas
        table = Table(data, colWidths=col_widths, repeatRows=1)
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ])
        table.setStyle(style)
        return table

    # Adicionar a primeira tabela ao PDF
    # Defina manualmente as larguras das colunas da primeira tabela
    col_widths_first_table = [1.3 * inch] * len(first_table_df.columns)  # Ajuste conforme necessário

    table1 = create_table_from_df(first_table_df, col_widths_first_table)
    elements.append(table1)
    elements.append(Spacer(1, 0.5 * inch))  # Espaço após a primeira tabela

    # Adicionar a segunda tabela ao PDF
    # Defina manualmente as larguras das colunas da segunda tabela
    col_widths_second_table = [2.8 * inch, 3.0 * inch, 2.0 * inch]  # Ajuste conforme necessário

    table2 = create_table_from_df(second_table_df, col_widths_second_table)
    elements.append(table2)
    elements.append(Spacer(1, 0.5 * inch))  # Espaço após a segunda tabela

    # Construir o documento PDF
    doc.build(elements)




def download_excel_pdf(file_xlsx, merged_pdf_file):
    # Criar um buffer de bytes para o arquivo ZIP
    zip_buffer = io.BytesIO()

    # Criar um arquivo ZIP contendo o Excel e o PDF mesclado
    with zipfile.ZipFile(zip_buffer, "w") as zip_file:
        # Adicionar o arquivo Excel ao ZIP
        zip_file.write(file_xlsx, arcname=file_xlsx)
        # Adicionar o arquivo PDF mesclado ao ZIP
        zip_file.write(merged_pdf_file, arcname=merged_pdf_file)

    # Mover o ponteiro para o início do buffer
    zip_buffer.seek(0)

    # Criar o botão de download para o arquivo ZIP contendo ambos os arquivos
    st.download_button(
        label="Download arquivos (Excel e PDF)",
        data=zip_buffer,
        file_name="arquivos.zip",
        mime="application/zip"
    )
